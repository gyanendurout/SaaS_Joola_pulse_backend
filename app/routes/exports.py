"""Export routes: JSON dump, CSV zip, Excel report, PDF report."""
from __future__ import annotations

import csv
import io
import json
import zipfile
from datetime import datetime
from uuid import UUID

from fastapi import APIRouter, HTTPException
from fastapi.responses import Response, StreamingResponse

from app.db import service_client

router = APIRouter(prefix="/api/runs/{run_id}", tags=["exports"])


def _fetch_run_bundle(run_id: str) -> dict:
    db = service_client()
    run = db.table("runs").select("*").eq("id", run_id).single().execute()
    if not run.data:
        raise HTTPException(404, "Run not found")
    pages = db.table("pages").select("*").eq("run_id", run_id).execute().data or []
    issues = db.table("issues").select("*").eq("run_id", run_id).execute().data or []
    entities = db.table("entities").select("*").eq("run_id", run_id).execute().data or []
    keywords = (
        db.table("keywords")
        .select("*")
        .eq("run_id", run_id)
        .order("search_volume", desc=True)
        .limit(200)
        .execute().data or []
    )
    competitors = (
        db.table("competitor_domains")
        .select("*")
        .eq("run_id", run_id)
        .order("intersections", desc=True)
        .execute().data or []
    )
    ranked = (
        db.table("domain_ranked_keywords")
        .select("*")
        .eq("run_id", run_id)
        .order("position")
        .limit(100)
        .execute().data or []
    )
    backlinks = (
        db.table("backlinks_summary")
        .select("*")
        .eq("run_id", run_id)
        .limit(1)
        .execute().data or []
    )

    recs_raw = run.data.get("recommendations")
    recs = {}
    if recs_raw:
        try:
            recs = json.loads(recs_raw) if isinstance(recs_raw, str) else recs_raw
        except Exception:
            pass

    return {
        "run": run.data,
        "pages": pages,
        "issues": issues,
        "entities": entities,
        "keywords": keywords,
        "competitors": competitors,
        "ranked_keywords": ranked,
        "backlinks": backlinks[0] if backlinks else None,
        "recommendations": recs,
    }


# ─── JSON ─────────────────────────────────────────────────────────────────────

@router.get("/export.json")
async def export_json(run_id: UUID):
    bundle = _fetch_run_bundle(str(run_id))
    body = json.dumps(bundle, default=str, indent=2)
    return Response(
        content=body,
        media_type="application/json",
        headers={"Content-Disposition": f'attachment; filename="seo-report-{run_id}.json"'},
    )


# ─── CSV zip ──────────────────────────────────────────────────────────────────

def _rows_to_csv(rows: list[dict]) -> bytes:
    if not rows:
        return b""
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=list(rows[0].keys()))
    writer.writeheader()
    for r in rows:
        writer.writerow({
            k: (json.dumps(v, default=str) if isinstance(v, (dict, list)) else v)
            for k, v in r.items()
        })
    return buf.getvalue().encode("utf-8")


@router.get("/export.csv")
async def export_csv(run_id: UUID):
    bundle = _fetch_run_bundle(str(run_id))
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        z.writestr("run.json", json.dumps(bundle["run"], default=str, indent=2))
        z.writestr("pages.csv", _rows_to_csv(bundle["pages"]))
        z.writestr("issues.csv", _rows_to_csv(bundle["issues"]))
        z.writestr("entities.csv", _rows_to_csv(bundle["entities"]))
        z.writestr("keywords.csv", _rows_to_csv(bundle["keywords"]))
        z.writestr("ranked_keywords.csv", _rows_to_csv(bundle["ranked_keywords"]))
        z.writestr("competitors.csv", _rows_to_csv(bundle["competitors"]))
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="seo-report-{run_id}.zip"'},
    )


# ─── Excel (.xlsx) ────────────────────────────────────────────────────────────

@router.get("/report.xlsx")
async def export_xlsx(run_id: UUID):
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        raise HTTPException(500, "openpyxl not installed. Run: pip install openpyxl")

    bundle = _fetch_run_bundle(str(run_id))
    run = bundle["run"]
    domain = run.get("canonical_domain", run.get("website_url", ""))
    generated = datetime.now().strftime("%Y-%m-%d %H:%M")

    wb = openpyxl.Workbook()

    # ── Styles ──
    HEADER_FILL = PatternFill("solid", fgColor="1E3A8A")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
    TITLE_FONT = Font(bold=True, size=13, color="1E3A8A")
    LABEL_FONT = Font(bold=True, size=10)
    ALT_FILL = PatternFill("solid", fgColor="F1F5F9")
    RED_FONT = Font(color="991B1B", bold=True)
    GREEN_FONT = Font(color="14532D", bold=True)

    def _header_row(ws, cols: list[str], row: int = 1):
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=row, column=ci, value=col)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 22

    def _auto_width(ws, min_w=10, max_w=50):
        for col in ws.columns:
            length = max(
                len(str(cell.value or "")) for cell in col
            )
            ws.column_dimensions[col[0].column_letter].width = min(max(length + 2, min_w), max_w)

    # ── Sheet 1: Overview ──
    ws = wb.active
    ws.title = "Overview"
    ws["A1"] = "SEO Intelligence Report"
    ws["A1"].font = Font(bold=True, size=16, color="1E3A8A")
    ws["A2"] = f"Domain: {domain}"
    ws["A2"].font = LABEL_FONT
    ws["A3"] = f"Market: {run.get('market', '')}  |  Language: {run.get('language', '')}"
    ws["A4"] = f"Analyzed: {run.get('created_at', '')}  |  Generated: {generated}"
    ws["A4"].font = Font(color="475569")
    ws.row_dimensions[1].height = 24

    kpis = [
        ("Issues (Critical)", sum(1 for i in bundle["issues"] if i.get("severity") == "critical")),
        ("Issues (High)", sum(1 for i in bundle["issues"] if i.get("severity") == "high")),
        ("Issues (Total)", len(bundle["issues"])),
        ("Keywords Found", len(bundle["keywords"])),
        ("Domain Currently Ranks For", len(bundle["ranked_keywords"])),
        ("Competitors Found", len(bundle["competitors"])),
        ("Domain Rank", (bundle["backlinks"] or {}).get("domain_rank", "N/A")),
        ("Referring Domains", (bundle["backlinks"] or {}).get("total_referring_domains", "N/A")),
        ("Total Backlinks", (bundle["backlinks"] or {}).get("total_backlinks", "N/A")),
    ]
    ws["A6"] = "Key Metrics"
    ws["A6"].font = TITLE_FONT
    _header_row(ws, ["Metric", "Value"], row=7)
    for ri, (label, val) in enumerate(kpis, 8):
        ws.cell(row=ri, column=1, value=label).font = LABEL_FONT
        ws.cell(row=ri, column=2, value=val)
        if ri % 2 == 0:
            ws.cell(row=ri, column=1).fill = ALT_FILL
            ws.cell(row=ri, column=2).fill = ALT_FILL
    _auto_width(ws)

    # ── Sheet 2: Issues ──
    ws2 = wb.create_sheet("SEO Issues")
    _header_row(ws2, ["Rule Code", "Severity", "Recommendation", "Details"])
    sev_order = {"critical": 0, "high": 1, "medium": 2, "low": 3, "info": 4}
    sorted_issues = sorted(bundle["issues"], key=lambda i: sev_order.get(i.get("severity", "info"), 5))
    for ri, iss in enumerate(sorted_issues, 2):
        ws2.cell(row=ri, column=1, value=iss.get("issue_code", ""))
        sev_cell = ws2.cell(row=ri, column=2, value=(iss.get("severity") or "").upper())
        sev = iss.get("severity", "")
        if sev == "critical":
            sev_cell.font = RED_FONT
        elif sev in ("high", "medium"):
            sev_cell.font = Font(color="92400E", bold=True)
        else:
            sev_cell.font = GREEN_FONT
        ws2.cell(row=ri, column=3, value=iss.get("recommendation", ""))
        ws2.cell(row=ri, column=4, value=json.dumps(iss.get("details") or {}, default=str))
        if ri % 2 == 0:
            for c in range(1, 5):
                ws2.cell(row=ri, column=c).fill = ALT_FILL
    _auto_width(ws2)

    # ── Sheet 3: Keywords ──
    ws3 = wb.create_sheet("Keywords")
    _header_row(ws3, ["Keyword", "Search Volume", "CPC ($)", "Difficulty", "Intent", "Type"])
    for ri, kw in enumerate(bundle["keywords"][:200], 2):
        ws3.cell(row=ri, column=1, value=kw.get("keyword", ""))
        ws3.cell(row=ri, column=2, value=kw.get("search_volume"))
        cpc = kw.get("cpc")
        ws3.cell(row=ri, column=3, value=round(float(cpc), 2) if cpc else None)
        ws3.cell(row=ri, column=4, value=kw.get("keyword_difficulty"))
        ws3.cell(row=ri, column=5, value=kw.get("intent", ""))
        ws3.cell(row=ri, column=6, value=kw.get("keyword_type", ""))
        if ri % 2 == 0:
            for c in range(1, 7):
                ws3.cell(row=ri, column=c).fill = ALT_FILL
    _auto_width(ws3)

    # ── Sheet 4: Recommendations ──
    ws4 = wb.create_sheet("Recommendations")
    recs_list = (bundle["recommendations"] or {}).get("recommendations") or []
    _header_row(ws4, ["Priority", "Where", "Title", "Currently", "Expected", "Benefit", "SEO Rule"])
    for ri, rec in enumerate(recs_list, 2):
        ws4.cell(row=ri, column=1, value=(rec.get("priority") or "").upper())
        ws4.cell(row=ri, column=2, value=rec.get("where", ""))
        ws4.cell(row=ri, column=3, value=rec.get("title", ""))
        ws4.cell(row=ri, column=4, value=rec.get("current", ""))
        ws4.cell(row=ri, column=5, value=rec.get("expected", rec.get("change", "")))
        ws4.cell(row=ri, column=6, value=rec.get("benefit", ""))
        ws4.cell(row=ri, column=7, value=rec.get("seo_rule", ""))
        if ri % 2 == 0:
            for c in range(1, 8):
                ws4.cell(row=ri, column=c).fill = ALT_FILL
    _auto_width(ws4)
    for cell in ws4["D"]:
        cell.alignment = Alignment(wrap_text=True)
    for cell in ws4["E"]:
        cell.alignment = Alignment(wrap_text=True)

    # ── Sheet 5: Rankings ──
    ws5 = wb.create_sheet("Rankings")
    _header_row(ws5, ["Position", "Keyword", "Search Volume", "CPC ($)", "Ranking URL"])
    for ri, kw in enumerate(bundle["ranked_keywords"], 2):
        pos = kw.get("position")
        pos_cell = ws5.cell(row=ri, column=1, value=pos)
        if pos and pos <= 3:
            pos_cell.font = GREEN_FONT
        elif pos and pos <= 10:
            pos_cell.font = Font(color="1E3A8A", bold=True)
        ws5.cell(row=ri, column=2, value=kw.get("keyword", ""))
        ws5.cell(row=ri, column=3, value=kw.get("search_volume"))
        cpc = kw.get("cpc")
        ws5.cell(row=ri, column=4, value=round(float(cpc), 2) if cpc else None)
        ws5.cell(row=ri, column=5, value=kw.get("url", ""))
        if ri % 2 == 0:
            for c in range(1, 6):
                ws5.cell(row=ri, column=c).fill = ALT_FILL
    _auto_width(ws5)

    # ── Sheet 6: Competitors ──
    ws6 = wb.create_sheet("Competitors")
    _header_row(ws6, ["#", "Domain", "Shared Keywords", "Avg Position"])
    for ri, comp in enumerate(bundle["competitors"], 2):
        ws6.cell(row=ri, column=1, value=ri - 1)
        ws6.cell(row=ri, column=2, value=comp.get("domain", ""))
        ws6.cell(row=ri, column=3, value=comp.get("intersections"))
        avg = comp.get("avg_position")
        ws6.cell(row=ri, column=4, value=round(float(avg), 1) if avg else None)
        if ri % 2 == 0:
            for c in range(1, 5):
                ws6.cell(row=ri, column=c).fill = ALT_FILL
    _auto_width(ws6)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"seo-report-{domain}-{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ─── PDF (print-ready HTML) ───────────────────────────────────────────────────

@router.get("/report.pdf")
async def export_pdf(run_id: UUID):
    """Generate a print-ready HTML report. Open in browser → Ctrl+P → Save as PDF."""
    bundle = _fetch_run_bundle(str(run_id))
    run = bundle["run"]
    domain = run.get("canonical_domain", run.get("website_url", ""))
    generated = datetime.now().strftime("%Y-%m-%d %H:%M")
    recs = bundle["recommendations"] or {}
    issues = bundle["issues"]
    keywords = bundle["keywords"][:50]
    ranked = bundle["ranked_keywords"][:30]
    comps = bundle["competitors"][:15]
    bl = bundle["backlinks"] or {}

    crit = sum(1 for i in issues if i.get("severity") == "critical")
    high = sum(1 for i in issues if i.get("severity") == "high")

    def _kw_rows() -> str:
        rows = ""
        for k in keywords:
            rows += (
                f"<tr>"
                f"<td>{k.get('keyword','')}</td>"
                f"<td>{k.get('search_volume','—')}</td>"
                f"<td>{('$'+str(round(float(k['cpc']),2))) if k.get('cpc') else '—'}</td>"
                f"<td>{k.get('keyword_difficulty','—')}</td>"
                f"<td>{k.get('intent','—')}</td>"
                f"</tr>"
            )
        return rows

    def _issue_rows() -> str:
        rows = ""
        sev_color = {"critical": "#991B1B", "high": "#9A3412", "medium": "#92400E", "low": "#475569"}
        for iss in issues:
            sev = iss.get("severity", "low")
            sev_c = sev_color.get(sev, "#000")
            rows += (
                f"<tr>"
                f"<td style='color:{sev_c};font-weight:700'>{sev.upper()}</td>"
                f"<td style='font-family:monospace'>{iss.get('issue_code','')}</td>"
                f"<td>{iss.get('recommendation','')}</td>"
                f"</tr>"
            )
        return rows

    def _rec_rows() -> str:
        rows = ""
        color = {"critical": "#991B1B", "high": "#9A3412", "medium": "#92400E", "low": "#475569"}
        for r in (recs.get("recommendations") or []):
            p = r.get("priority", "low")
            p_c = color.get(p, "#000")
            rows += (
                f"<tr>"
                f"<td style='color:{p_c};font-weight:700'>{p.upper()}</td>"
                f"<td><strong>{r.get('title','')}</strong></td>"
                f"<td>{r.get('current','—')}</td>"
                f"<td>{r.get('expected',r.get('change','—'))}</td>"
                f"<td>{r.get('benefit','—')}</td>"
                f"</tr>"
            )
        return rows

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>SEO Report — {domain}</title>
<style>
  @page {{ size: A4; margin: 18mm 15mm; }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: 'Segoe UI', Arial, sans-serif; font-size: 10pt; color: #0F172A; background: #fff; }}
  h1 {{ font-size: 22pt; color: #1E3A8A; margin-bottom: 4px; }}
  h2 {{ font-size: 13pt; color: #1E3A8A; margin: 20px 0 8px; border-bottom: 2px solid #1E3A8A; padding-bottom: 4px; }}
  h3 {{ font-size: 11pt; color: #1E3A8A; margin: 12px 0 6px; }}
  .meta {{ color: #475569; font-size: 9pt; margin-bottom: 20px; }}
  .kpis {{ display: grid; grid-template-columns: repeat(3, 1fr); gap: 10px; margin-bottom: 20px; }}
  .kpi {{ border: 1px solid #E2E8F0; border-radius: 6px; padding: 10px 14px; }}
  .kpi-val {{ font-size: 20pt; font-weight: 800; color: #1E3A8A; }}
  .kpi-label {{ font-size: 8.5pt; color: #475569; font-weight: 600; text-transform: uppercase; letter-spacing: 0.05em; }}
  table {{ width: 100%; border-collapse: collapse; margin-bottom: 16px; font-size: 9pt; }}
  th {{ background: #1E3A8A; color: #fff; padding: 6px 8px; text-align: left; font-size: 8.5pt; }}
  td {{ padding: 5px 8px; border-bottom: 1px solid #E2E8F0; }}
  tr:nth-child(even) td {{ background: #F8FAFC; }}
  .summary-box {{ background: #EFF6FF; border: 1px solid #BFDBFE; border-radius: 6px; padding: 12px 16px; margin-bottom: 18px; }}
  .page-break {{ page-break-before: always; }}
  @media print {{
    body {{ -webkit-print-color-adjust: exact; print-color-adjust: exact; }}
    .no-print {{ display: none; }}
  }}
</style>
</head>
<body>
<div class="no-print" style="background:#1E3A8A;color:#fff;padding:12px 20px;text-align:center;font-size:13px">
  Press <strong>Ctrl+P</strong> (or Cmd+P on Mac) → Change destination to <strong>Save as PDF</strong> → Save
</div>

<h1>SEO Intelligence Report</h1>
<p class="meta">Domain: <strong>{domain}</strong> &nbsp;|&nbsp; Market: {run.get('market','')} &nbsp;|&nbsp;
Analyzed: {run.get('created_at','')[:10]} &nbsp;|&nbsp; Generated: {generated}</p>

{ f'<div class="summary-box"><strong>AI Summary:</strong> {recs.get("executive_summary","")}</div>' if recs.get("executive_summary") else "" }

<h2>Key Performance Indicators</h2>
<div class="kpis">
  <div class="kpi"><div class="kpi-val" style="color:{'#991B1B' if crit>0 else '#14532D'}">{crit}</div><div class="kpi-label">Critical Issues</div></div>
  <div class="kpi"><div class="kpi-val" style="color:{'#9A3412' if high>0 else '#14532D'}">{high}</div><div class="kpi-label">High Issues</div></div>
  <div class="kpi"><div class="kpi-val">{len(issues)}</div><div class="kpi-label">Total Issues</div></div>
  <div class="kpi"><div class="kpi-val">{len(keywords)}</div><div class="kpi-label">Keywords Found</div></div>
  <div class="kpi"><div class="kpi-val">{len(ranked)}</div><div class="kpi-label">Currently Ranking</div></div>
  <div class="kpi"><div class="kpi-val">{bl.get('domain_rank','—')}</div><div class="kpi-label">Domain Rank</div></div>
  <div class="kpi"><div class="kpi-val">{str(bl.get('total_referring_domains','—'))}</div><div class="kpi-label">Referring Domains</div></div>
  <div class="kpi"><div class="kpi-val">{len(comps)}</div><div class="kpi-label">Competitors</div></div>
</div>

<h2>SEO Issues</h2>
<table>
  <thead><tr><th>Severity</th><th>Rule</th><th>Recommendation</th></tr></thead>
  <tbody>{_issue_rows()}</tbody>
</table>

<div class="page-break"></div>
<h2>Recommendations</h2>
<table>
  <thead><tr><th>Priority</th><th>Action</th><th>Currently</th><th>Expected</th><th>Benefit</th></tr></thead>
  <tbody>{_rec_rows()}</tbody>
</table>

<h2>Keyword Opportunities (Top 50)</h2>
<table>
  <thead><tr><th>Keyword</th><th>Volume</th><th>CPC</th><th>Difficulty</th><th>Intent</th></tr></thead>
  <tbody>{_kw_rows()}</tbody>
</table>

<div class="page-break"></div>
<h2>Competitors</h2>
<table>
  <thead><tr><th>#</th><th>Domain</th><th>Shared Keywords</th><th>Avg Position</th></tr></thead>
  <tbody>
    {"".join(f"<tr><td>{i+1}</td><td>{c.get('domain','')}</td><td>{c.get('intersections','—')}</td><td>{round(float(c['avg_position']),1) if c.get('avg_position') else '—'}</td></tr>" for i,c in enumerate(comps))}
  </tbody>
</table>

<h2>Current Rankings (Top 30)</h2>
<table>
  <thead><tr><th>Position</th><th>Keyword</th><th>Volume</th><th>URL</th></tr></thead>
  <tbody>
    {"".join(f"<tr><td><strong>#{k.get('position','—')}</strong></td><td>{k.get('keyword','')}</td><td>{k.get('search_volume','—')}</td><td style='font-size:8pt'>{(k.get('url') or '')[:60]}</td></tr>" for k in ranked)}
  </tbody>
</table>

<p style="margin-top:30px;color:#94A3B8;font-size:8.5pt;text-align:center">
  AI SEO Manager — {domain} — {generated}
</p>
</body>
</html>"""

    return Response(
        content=html,
        media_type="text/html",
        headers={"Content-Disposition": f'inline; filename="seo-report-{domain}.html"'},
    )
